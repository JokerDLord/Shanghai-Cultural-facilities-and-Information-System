import openpyxl
from openpyxl import Workbook
import os
import math


def write_en(select_lst):
    features=[0 for i in range(16)]
    for i in range(1,17):
        if str(i) in select_lst:
            features[i-1]=1
    return features

def write_tag(select_lst):
    select_lst=str(select_lst)
    features=[0 for i in range(12)]
    for j in range(12):
        if str(j+1) in select_lst:
            features[j]=1
    return features

def write_place(select_lst):
    features=[0 for i in range(14)]
    places=['电影院','公园','KTV','博物馆','天文馆','美术馆',\
    '水族馆','纪念馆','展览馆','图书馆','科技馆','商场',\
    '人文旅游景点（名胜古迹）','自然旅游景点']
    for i in range(14):
        if places[i] in select_lst:
            features[i]=1
    return features

def write_title(sheet):
    ttitle=['gender','age','num','transport']
    enter=['enter_'+str(i) for i in range(1,17)]
    tag=['tag_'+str(i) for i in range(1,13)]
    place=['place_'+str(i) for i in range(1,15)]
    ttitle.extend(enter)
    ttitle.extend(tag)
    ttitle.extend(place)
    for i in range(len(ttitle)):
        _=sheet.cell(row=1,column=i+1,value=ttitle[i])
    
def write_row(sheet,row):
    row1=sheet.max_row+1
    gender=row[1].value
    age=row[2].value
    num=row[3].value
    transport=row[4].value
    enter_lst=row[5].value
    enter=write_en(enter_lst)
    tag_lst=row[6].value
    tag=write_tag(tag_lst)
    place_lst=row[7].value
    place=write_place(place_lst)
    write_lst=[gender,age,num,transport]
    write_lst.extend(enter)
    write_lst.extend(tag)
    write_lst.extend(place)
    for i in range(len(write_lst)):
        _=sheet.cell(row=row1,column=i+1,value=write_lst[i])

def write(inpath,outpath):
    wb=Workbook()
    sheet=wb.active
    write_title(sheet)
    st=openpyxl.load_workbook(inpath).active
    for row in st[2:st.max_row]:
        write_row(sheet,row)
    wb.save(outpath)


inputpath='人群娱乐出行选择调查.xlsx'
outpath='searchtable.xlsx'
write(inputpath,outpath)

